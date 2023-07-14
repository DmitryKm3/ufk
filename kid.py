#!/home/ufk/ufk_venv/bin/python
import django
import os
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "monitoring_project.settings")
django.setup()

from datetime import datetime, timezone
import pytz
localtime = pytz.timezone("Asia/Krasnoyarsk")

from monitoring.models import UK, PIF
from django.db.models import Q

import time
import re
import shutil
import openpyxl
#smtp import
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication


#функция kidpy принимает в качестве аргументов дату и параметры выборки из post запроса
#при вызове функции запускается скрипт формирования отчета
import importlib
 #функция будет выполняться асинхронно, чтобы не положить работы сайта, хотя может он и так не ляжет, я не проверял:D
def kidpy(DateFromPost, ParamertsFromPost):

    def filter_out(content, rules, count=1):
        filtered_content = []
        rules_list = rules.split(';;;')
        for section in content:
            if any([re.match(rule, section[0], re.IGNORECASE) for rule in rules_list]):
                filtered_content += section[1]
                continue
            filtered_content += [doc for doc in section[1] if any([re.match(rule, doc[0], re.IGNORECASE) for rule in rules_list])]
        return sorted(filtered_content, key=lambda tup: tup[1], reverse=True)[:count]

    def write_line(ws, row, data):
        try:
            npp = ws.cell(row=row-1, column=1).value + 1
        except:
            npp = 1
        ws.cell(row=row, column=1).value = npp
        for i in range(6):
            ws.cell(row=row, column=2+i).value = data[i]

    #uks = UK.objects.filter(~Q(uk_extractor = ''))
    uks = UK.objects.all()
    sum = 0
    start_time = time.time()
    report_data = []
    total_start = time.time()
    localtime = pytz.timezone("Asia/Krasnoyarsk")
    date_now = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=localtime)
    date_str = datetime.strftime(date_now, '%Y%m%d')
    filename = f'static/report_kid_{date_str}.xlsx'
    shutil.copy('static/templates/report_kid.xlsx', filename)
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    wb.save(filename=filename)
    for uk in uks:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        current_row = ws.max_row + 1
        print(f'Doing UK: {uk.uk_name}')
        start_time = time.time()
        module = None

        if uk.uk_extractor:
           module = importlib.import_module(f'extractors.{uk.uk_extractor}')
    #      for pif in PIF.objects.filter(Q(pif_uk = uk) & ~Q(pif_checkpage='')):

        for pif in PIF.objects.filter(Q(pif_uk = uk)):
           ext = None
           print(pif.pif_name)
           if not pif.pif_enabled:
               continue
           if pif.pif_checkpage and module:
               result = []
               try:
                   ext = module.Extractor(pif.pif_checkpage)
                   ext.scrape()
                   content = filter_out(ext.get_data(DateFromPost), ParamertsFromPost, 7)
                   for doc in content:
                       report_data.append((pif.pif_npp, pif.pif_name, '://'.join((uk.uk_sitetype, uk.uk_site)), \
                               doc[0], doc[1].strftime("%d.%m.%Y %H:%M"), doc[2]))
                       result = [pif.pif_uk.uk_name, pif.pif_name, pif.pif_checkpage, doc[0], doc[2], doc[1].strftime("%d.%m.%Y %H:%M")]
                       write_line(ws, current_row, result)
                       current_row = current_row + 1
               except:
                   report_data.append((pif.pif_npp, pif.pif_name, '://'.join((uk.uk_sitetype, uk.uk_site)), \
                       "", "", ""))
                   print(f'Ошибка при получении: {pif.pif_name} \t {pif.pif_checkpage}')
               if len(result) == 0:
                   result = [pif.pif_uk.uk_name, pif.pif_name, pif.pif_checkpage, '', '', '']
                   write_line(ws, current_row, result)
                   current_row = current_row + 1
               else:
                   report_data.append((pif.pif_npp, pif.pif_name, '://'.join((uk.uk_sitetype, uk.uk_site)), \
                           "", "", ""))

               if pif.pif_checkpage.strip() == '':
                   result = [pif.pif_uk.uk_name, pif.pif_name, pif.pif_checkpage, '', '', '']
                   write_line(ws, current_row, result)
                   current_row = current_row + 1

               if len(result) == 0:
                   result = [pif.pif_uk.uk_name, pif.pif_name, pif.pif_checkpage, '', '', '']
                   write_line(ws, current_row, result)
                   current_row = current_row + 1

               if module and ext:
                   del ext
           wb.save(filename)
           del module
           print(f'UK execution time: {time.time() - start_time}')
           #for row in report_data:
           #    ws.append(row)
           #wb.save(filename=filename)
           print(f'Total execution time: {time.time() - total_start}')



def send_report_email(to_email):
     # Настройки SMTP-сервера
     smtp_server = 'smtp.gmail.com'
     smtp_port = 587
     smtp_username = 'otchet064@gmail.com'
     smtp_password = 'vbgzqbkjukpuppkv'
     # Создание объекта MIMEMultipart для формирования письма
     msg = MIMEMultipart()
     msg['From'] = smtp_username
     msg['To'] = to_email
     date_now = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=localtime)
     date_str = datetime.strftime(date_now, '%Y%m%d')
     filename = f'static/report_kid_{date_str}.xlsx'
     attachment_path = filename
     # Добавление текста письма
     msg.attach(MIMEText('Отчет'))
     # Создание объекта MIMEApplication для вложения файла
     with open(attachment_path, 'rb') as file:
         attachment = MIMEApplication(file.read(), Name='report.xlsx')
     # Добавление заголовков для вложения
     attachment['Content-Disposition'] = f'attachment; filename="{attachment_path}"'
     # Добавление вложения к письму
     msg.attach(attachment)
     print("отправил")
     # Создание соединения с SMTP-сервером
     with smtplib.SMTP(smtp_server, smtp_port) as server:
         server.starttls()
         server.login(smtp_username, smtp_password)
         server.send_message(msg)
